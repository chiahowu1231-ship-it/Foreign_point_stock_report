
# src/run_report.py
# 純外資整合版（a 代號各異 / b 先試 b=a，必要時 fallback b=9900）
# ------------------------------------------------------------
# 主要功能（整合版）：
# - 每家外資內部：依「淨超」由大到小排序
# - 外資排序：依「各外資總淨超」由大到小排序（A1）
# - summary.json：加入 top_preview（每家 Top 10）供 mailer 直接引用（A2）
# - Excel：新增 TopByBroker sheet（每家 Top 10）（A3）
# - DEBUG_HTML=1：輸出 output/debug/*.html（E/B 各一份）
# - 強化 parse_table：不只 script，也會從 href/onclick/tr html 取 GenLink2stk
# - yfinance 噪音抑制 + 快取：不再滿版 404/possibly delisted
# - exit code 規則更合理：
#     * total_rows==0 或 brokers_ok==0 才 exit 1
#     * 有資料（total_rows>0）即使 brokers_fail>0 也 exit 0

import os
import re
import json
import time
import random
from datetime import datetime
from zoneinfo import ZoneInfo
import contextlib
import io

import pandas as pd
import requests
from bs4 import BeautifulSoup
import yfinance as yf

# 你的 src/config.py 必須是 tuple 結構：
# LEGEND_BROKERS = { "1470_F": ("1470","label"), ... }
from config import LEGEND_BROKERS

TZ = ZoneInfo("Asia/Taipei")
DEFAULT_DAYS = 5
BASE_URL = "https://fubon-ebrokerdj.fbs.com.tw/z/zg/zgb/zgb0.djhtm"

UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

NO_DATA_TEXT = "無此券商分點交易資料"
PRICE_CACHE = {}  # 現價快取（避免同股票重複查）


def ensure_output_dir():
    os.makedirs("output", exist_ok=True)


def now_taipei():
    return datetime.now(TZ)


def safe_int(s: str) -> int:
    if s is None:
        return 0
    s = s.strip().replace(",", "")
    if s in ("", "--"):
        return 0
    try:
        return int(s)
    except ValueError:
        m = re.search(r"-?\d+", s)
        return int(m.group()) if m else 0


def dump_debug_html(tag: str, html: str):
    """DEBUG_HTML=1 時，把抓回來的 HTML 存到 output/debug/"""
    if os.getenv("DEBUG_HTML", "0") != "1":
        return
    os.makedirs("output/debug", exist_ok=True)
    path = os.path.join("output", "debug", f"{tag}.html")
    with open(path, "w", encoding="utf-8", errors="ignore") as f:
        f.write(html)


def fetch_html(url: str, timeout: int = 20, retries: int = 5, base_sleep: float = 1.0) -> str:
    """指數退避重試 + jitter（確實重試到 retries 次）"""
    headers = {
        "User-Agent": UA,
        "Referer": "https://fubon-ebrokerdj.fbs.com.tw/",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.8",
        "Connection": "keep-alive",
    }

    last_err = None
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=headers, timeout=timeout)
            r.encoding = "big5"
            if r.status_code == 200 and r.text:
                return r.text
            last_err = RuntimeError(f"HTTP {r.status_code}")
        except Exception as e:
            last_err = e

        time.sleep(base_sleep * (2 ** attempt) + random.uniform(0, 0.5))

    raise last_err


def parse_table(html: str) -> dict:
    """
    強化解析：
    - 從 script / a[href|onclick] / tr html 找 GenLink2stk('ASxxxx','name')
    - 支援單/雙引號與空白換行
    - 欄位優先找 class t3n0/t3n1，不足則 fallback 所有 td
    """
    soup = BeautifulSoup(html, "html.parser")
    res = {}

    pat = re.compile(
        r"GenLink2stk\(\s*'AS(\w+)'\s*,\s*'(.+?)'\s*\)"
        r"|GenLink2stk\(\s*\"AS(\w+)\"\s*,\s*\"(.+?)\"\s*\)"
    )

    for tr in soup.find_all("tr"):
        found = None

        # ① script
        for sc in tr.find_all("script"):
            txt = sc.get_text() or ""
            m = pat.search(txt)
            if m:
                sid = m.group(1) or m.group(3)
                name = m.group(2) or m.group(4)
                found = (sid, name)
                break

        # ② href / onclick
        if not found:
            for a in tr.find_all("a"):
                blob = (a.get("href", "") or "") + " " + (a.get("onclick", "") or "")
                m = pat.search(blob)
                if m:
                    sid = m.group(1) or m.group(3)
                    name = m.group(2) or m.group(4)
                    found = (sid, name)
                    break

        # ③ tr html fallback
        if not found:
            m = pat.search(str(tr))
            if m:
                sid = m.group(1) or m.group(3)
                name = m.group(2) or m.group(4)
                found = (sid, name)

        if not found:
            continue

        sid, name = found

        tds = tr.find_all("td", class_=["t3n1", "t3n0"])
        if len(tds) < 3:
            tds = tr.find_all("td")
        if len(tds) < 3:
            continue

        res[sid] = {
            "name": name,
            "buy": tds[0].get_text(strip=True),
            "sell": tds[1].get_text(strip=True),
            "net": safe_int(tds[2].get_text(strip=True)),
        }

    return res


def get_stock_price(sid: str) -> float:
    """
    靜音 yfinance 的 404/possibly delisted 噪音 + 快取
    抓不到就回 0.0（不影響報表產出）
    """
    if sid in PRICE_CACHE:
        return PRICE_CACHE[sid]

    buf = io.StringIO()
    price = 0.0

    for suffix in [".TW", ".TWO"]:
        try:
            with contextlib.redirect_stdout(buf), contextlib.redirect_stderr(buf):
                data = yf.Ticker(f"{sid}{suffix}").history(period="1d")
            if data is not None and not data.empty:
                price = float(data["Close"].iloc[-1])
                break
        except Exception:
            continue

    PRICE_CACHE[sid] = price
    return price


def try_fetch_and_parse(a_code: str, b_code: str, days: int):
    """抓 E/B 並解析；回傳 (qty_map, amt_map, url_qty, url_amt, no_data_page)"""
    url_qty = f"{BASE_URL}?a={a_code}&b={b_code}&c=E&d={days}"
    url_amt = f"{BASE_URL}?a={a_code}&b={b_code}&c=B&d={days}"

    html_qty = fetch_html(url_qty)
    html_amt = fetch_html(url_amt)

    dump_debug_html(f"{a_code}_{b_code}_E", html_qty)
    dump_debug_html(f"{a_code}_{b_code}_B", html_amt)

    if (NO_DATA_TEXT in html_qty) and (NO_DATA_TEXT in html_amt):
        return {}, {}, url_qty, url_amt, True

    qty_map = parse_table(html_qty)
    amt_map = parse_table(html_amt)

    return qty_map, amt_map, url_qty, url_amt, False


def build_report(days: int):
    start_time = now_taipei()

    rows = []
    failures = []
    errors = []

    broker_ok = 0
    broker_fail = 0

    for broker_key, (a_code, broker_label) in LEGEND_BROKERS.items():
        a_code = str(a_code)

        # 先試 b=a（你已驗證 1470/1360 都是 b=a 有資料）
        # 若你確定全部都 b=a，可改成 candidate_bs = [a_code] 會更快
        candidate_bs = [a_code, "9900"]

        chosen_b = None
        qty_map = amt_map = {}
        url_qty = url_amt = ""

        try:
            for b_try in candidate_bs:
                q_map, a_map, u_qty, u_amt, no_data = try_fetch_and_parse(a_code, b_try, days)
                url_qty, url_amt = u_qty, u_amt

                if no_data:
                    continue

                if q_map or a_map:
                    chosen_b = b_try
                    qty_map, amt_map = q_map, a_map
                    break

            if not chosen_b:
                raise RuntimeError(f"查詢結果：{NO_DATA_TEXT}（已嘗試 b=a 與 b=9900）")

            common = set(qty_map.keys()) & set(amt_map.keys())
            if not common:
                raise RuntimeError("E/B 解析後無交集（欄位定位或解析規則可能需調整）")

            for sid in common:
                info = qty_map[sid]
                net_qty = info["net"]
                net_amt = amt_map[sid]["net"]  # 仟元
                avg_cost = round((net_amt / net_qty), 2) if net_qty > 0 else 0.0

                price = get_stock_price(sid)
                bias = (
                    round(((price - avg_cost) / avg_cost) * 100, 2)
                    if (price > 0 and avg_cost > 0)
                    else 0.0
                )

                rows.append({
                    "日期": start_time.strftime("%Y%m%d"),
                    "代碼": sid,
                    "名稱": info["name"],
                    "大戶": broker_label,
                    "買進": info["buy"],
                    "賣出": info["sell"],
                    "淨超": int(net_qty),
                    "區間均價": avg_cost,
                    "現價": round(price, 2),
                    "乖離率": f"{bias}%",
                })

            broker_ok += 1
            time.sleep(1.2)

        except Exception as e:
            broker_fail += 1
            msg = f"{broker_label}({broker_key}) 失敗：{str(e)}"
            errors.append(msg)
            failures.append({
                "日期": start_time.strftime("%Y%m%d"),
                "券商key": str(broker_key),
                "總公司(a)": a_code,
                "嘗試分點(b)": ",".join(candidate_bs),
                "券商名稱": broker_label,
                "錯誤訊息": str(e),
                "網址(張數E)": url_qty,
                "網址(金額B)": url_amt,
            })

    df = pd.DataFrame(rows, columns=[
        "日期", "代碼", "名稱", "大戶", "買進", "賣出", "淨超", "區間均價", "現價", "乖離率"
    ])

    fail_df = pd.DataFrame(failures, columns=[
        "日期", "券商key", "總公司(a)", "嘗試分點(b)", "券商名稱", "錯誤訊息", "網址(張數E)", "網址(金額B)"
    ])

    summary = {
        "generated_at": start_time.isoformat(),
        "timezone": "Asia/Taipei",
        "days": days,
        "total_rows": int(len(df)),
        "brokers_total": int(len(LEGEND_BROKERS)),
        "brokers_ok": broker_ok,
        "brokers_fail": broker_fail,
        # success 定義：全部券商成功才 True（用於 email 顯示狀態）
        "success": (broker_fail == 0),
        "errors": errors[:50],
    }

    # ===== A1：外資排序 + 每家內部淨超排序 =====
    if not df.empty:
        df["淨超"] = pd.to_numeric(df["淨超"], errors="coerce").fillna(0).astype(int)

        broker_order = (
            df.groupby("大戶")["淨超"]
              .sum()
              .sort_values(ascending=False)
              .index
              .tolist()
        )
        df["大戶"] = pd.Categorical(df["大戶"], categories=broker_order, ordered=True)
        df = df.sort_values(["大戶", "淨超"], ascending=[True, False]).reset_index(drop=True)

        # 把排序後 df 回存（讓後續 export_excel 使用的 df 已排序）
        # 注意：df 在此作用域內是局部變數，下面 return 會帶出去

    # ===== A2：summary.json 加入每家外資 Top N（Top 10） =====
    TOP_N = int(os.getenv("TOP_N", "10"))  # 你指定要 10
    top_preview = []

    if not df.empty and isinstance(df.get("大戶").dtype, pd.CategoricalDtype):
        for broker in df["大戶"].cat.categories.tolist():
            sub = df[df["大戶"] == broker].head(TOP_N)
            if sub.empty:
                continue

            top_preview.append({
                "broker": str(broker),
                "total_net": int(df[df["大戶"] == broker]["淨超"].sum()),
                "rows": [
                    {
                        "sid": r["代碼"],
                        "name": r["名稱"],
                        "net": int(r["淨超"]),
                        "avg": r["區間均價"],
                        "price": r["現價"],
                        "bias": r["乖離率"],
                    }
                    for _, r in sub.iterrows()
                ]
            })

    summary["top_preview"] = top_preview
    summary["top_n"] = TOP_N

    # 若總筆數 0 → success=false（避免假成功）
    if summary["total_rows"] == 0:
        summary["success"] = False
        summary["errors"] = (summary.get("errors") or []) + [
            "總資料筆數為 0（全部查無分點交易資料或解析規則需調整）"
        ]

    return df, fail_df, summary


def export_excel(df: pd.DataFrame, fail_df: pd.DataFrame, xlsx_path: str):
    from openpyxl.styles import PatternFill, Font
    from openpyxl.formatting.rule import FormulaRule

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Report")
        ws = writer.sheets["Report"]

        col_widths = {
            "A": 10, "B": 8, "C": 14, "D": 40,
            "E": 10, "F": 10, "G": 10, "H": 12,
            "I": 10, "J": 10
        }
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w

        # 乖離率(J欄)條件式上色：淺綠 / 藍 / 深紅
        last_row = ws.max_row
        if last_row >= 2:
            rng = f"J2:J{last_row}"

            fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            fill_blue  = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
            fill_red   = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
            font_white = Font(color="FFFFFF", bold=True)

            rule_red = FormulaRule(
                formula=[r'=VALUE(SUBSTITUTE($J2,"%",""))>10'],
                fill=fill_red, font=font_white, stopIfTrue=True
            )
            rule_blue = FormulaRule(
                formula=[r'=AND(VALUE(SUBSTITUTE($J2,"%",""))>3,VALUE(SUBSTITUTE($J2,"%",""))<=10)'],
                fill=fill_blue, stopIfTrue=True
            )
            rule_green = FormulaRule(
                formula=[r'=VALUE(SUBSTITUTE($J2,"%",""))<=3'],
                fill=fill_green, stopIfTrue=True
            )

            ws.conditional_formatting.add(rng, rule_red)
            ws.conditional_formatting.add(rng, rule_blue)
            ws.conditional_formatting.add(rng, rule_green)

        # ===== A3：TopByBroker sheet（每家外資 Top 10）=====
        top_n = int(os.getenv("TOP_N", "10"))
        if df is not None and not df.empty:
            top_df = df.groupby("大戶", sort=False).head(top_n)
            top_df.to_excel(writer, index=False, sheet_name="TopByBroker")
            ws3 = writer.sheets["TopByBroker"]
            for col, w in {
                "A": 10, "B": 8, "C": 14, "D": 40,
                "E": 10, "F": 10, "G": 10, "H": 12,
                "I": 10, "J": 10
            }.items():
                ws3.column_dimensions[col].width = w

        if fail_df is not None and not fail_df.empty:
            fail_df.to_excel(writer, index=False, sheet_name="Failures")
            ws2 = writer.sheets["Failures"]
            ws2.column_dimensions["A"].width = 10
            ws2.column_dimensions["B"].width = 12
            ws2.column_dimensions["C"].width = 10
            ws2.column_dimensions["D"].width = 14
            ws2.column_dimensions["E"].width = 35
            ws2.column_dimensions["F"].width = 60
            ws2.column_dimensions["G"].width = 55
            ws2.column_dimensions["H"].width = 55


def export_pdf(df: pd.DataFrame, pdf_path: str, summary: dict):
    """
    優先用嵌入字型（fonts/NotoSansTC-Regular.ttf）避免亂碼；
    若字型缺失，仍產出一份「提示 PDF」避免 artifacts 找不到。
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_path = os.path.join("fonts", "NotoSansTC-Regular.ttf")
    font_name = "NotoSansTC"

    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=landscape(A4),
        leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18
    )

    styles = getSampleStyleSheet()
    normal = styles["Normal"]
    title_style = styles["Title"]

    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont(font_name, font_path))
        normal.fontName = font_name
        title_style.fontName = font_name
    else:
        normal.fontName = "Helvetica"
        title_style.fontName = "Helvetica"

    title = Paragraph("【外資分點狙擊分析】報表", title_style)
    meta = Paragraph(
        f"產生時間：{summary['generated_at']}（{summary['timezone']}）　"
        f"查詢區間：{summary['days']} 日　"
        f"結果：{'成功' if summary['success'] else '失敗/部分失敗'}　"
        f"筆數：{summary['total_rows']}",
        normal
    )
    elements = [title, Spacer(1, 8), meta, Spacer(1, 12)]

    if not os.path.exists(font_path):
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("⚠️ 缺少字型 fonts/NotoSansTC-Regular.ttf（PDF 以 Helvetica 產生）", normal))

    header = list(df.columns) if df is not None and not df.empty else ["日期","代碼","名稱","大戶","買進","賣出","淨超","區間均價","現價","乖離率"]
    data = [header] + (df.astype(str).values.tolist() if df is not None and not df.empty else [])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), normal.fontName),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EAEAEA")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F7F7")]),
    ]))
    elements.append(table)

    if summary.get("errors"):
        elements.append(Spacer(1, 12))
        elements.append(Paragraph("錯誤摘要（前 20 筆）：", normal))
        for e in summary["errors"][:20]:
            elements.append(Paragraph(f"- {e}", normal))

    doc.build(elements)


def main():
    ensure_output_dir()
    days = int(os.getenv("DAYS", str(DEFAULT_DAYS)))

    df, fail_df, summary = build_report(days)

    ymd = now_taipei().strftime("%Y%m%d")
    xlsx_path = os.path.join("output", f"IKE_Report_{ymd}.xlsx")
    pdf_path = os.path.join("output", f"IKE_Report_{ymd}.pdf")
    summary_path = os.path.join("output", "summary.json")

    export_excel(df, fail_df, xlsx_path)
    export_pdf(df, pdf_path, summary)

    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    print(f"[OK] Excel: {xlsx_path}")
    print(f"[OK] PDF  : {pdf_path}")
    print(f"[OK] Summary: {summary_path}")

    # ✅ exit code 規則（整合版）
    # 只有在「完全沒資料」或「全部券商都失敗」才 exit 1
    if summary.get("total_rows", 0) == 0 or summary.get("brokers_ok", 0) == 0:
        raise SystemExit(1)

    # 只要有資料（total_rows>0），即使 brokers_fail>0 也 exit 0
    return


if __name__ == "__main__":
    main()
